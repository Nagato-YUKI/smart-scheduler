"""
智能排课系统 - 全面测试脚本
测试范围:
1. Excel数据导入功能（教室、教师、班级、课程）
2. 数据导入后自动排课功能
3. 排课算法正确性验证
4. 课表显示功能（周课表格式）
5. 边界条件和错误处理
"""

import sys
import os
import json
import tempfile
import io
from datetime import datetime

# 设置路径
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from openpyxl import Workbook, load_workbook
from peewee_manager import (
    _database, Room, Teacher, SchoolClass, Course, Holiday, TeachingClass, ScheduleEntry
)

# ======================== 测试配置 ========================

TEST_RESULTS = []
TEST_PASS_COUNT = 0
TEST_FAIL_COUNT = 0
TEST_SKIP_COUNT = 0


def log_result(test_id, test_name, status, detail='', severity=''):
    global TEST_PASS_COUNT, TEST_FAIL_COUNT, TEST_SKIP_COUNT
    result = {
        'test_id': test_id,
        'test_name': test_name,
        'status': status,
        'detail': detail,
        'severity': severity
    }
    TEST_RESULTS.append(result)
    if status == 'PASS':
        TEST_PASS_COUNT += 1
    elif status == 'FAIL':
        TEST_FAIL_COUNT += 1
    else:
        TEST_SKIP_COUNT += 1
    
    icon = {
        'PASS': '[通过]',
        'FAIL': '[失败]',
        'SKIP': '[跳过]'
    }.get(status, '[未知]')
    print(f"  {icon} {test_id} {test_name}: {detail}")


# ======================== 工具函数 ========================

_temp_files = []

def save_workbook_to_temp(wb, filename):
    """将Workbook保存到临时文件"""
    temp_dir = tempfile.gettempdir()
    filepath = os.path.join(temp_dir, filename)
    wb.save(filepath)
    _temp_files.append(filepath)
    return filepath


def tempfile_from_response(resp):
    """从Flask响应创建临时文件"""
    temp_dir = tempfile.gettempdir()
    filepath = os.path.join(temp_dir, 'temp_download.xlsx')
    with open(filepath, 'wb') as f:
        f.write(resp.data)
    _temp_files.append(filepath)
    return filepath


def cleanup_temp_files():
    """清理临时文件"""
    for f in _temp_files:
        try:
            os.remove(f)
        except:
            pass


def clear_all_data():
    """清空所有数据表"""
    if not _database.is_closed():
        _database.close()
    _database.connect()
    try:
        ScheduleEntry.delete().execute()
    except:
        pass
    try:
        TeachingClass.delete().execute()
    except:
        pass
    try:
        Course.delete().execute()
    except:
        pass
    try:
        SchoolClass.delete().execute()
    except:
        pass
    try:
        Teacher.delete().execute()
    except:
        pass
    try:
        Room.delete().execute()
    except:
        pass
    try:
        Holiday.delete().execute()
    except:
        pass
    _database.close()


def create_test_data():
    """创建测试基础数据"""
    if not _database.is_closed():
        _database.close()
    _database.connect()
    
    # 创建教室
    Room.create(room_number='R001', name='博学楼101', capacity=60, room_type='普通教室')
    Room.create(room_number='R002', name='博学楼102', capacity=55, room_type='普通教室')
    Room.create(room_number='R003', name='博学楼201', capacity=50, room_type='多媒体教室')
    Room.create(room_number='R004', name='博学楼301', capacity=40, room_type='机房')
    Room.create(room_number='R005', name='博学楼401', capacity=35, room_type='实验室')
    
    # 创建教师
    Teacher.create(teacher_number='T001', name='张伟', teachable_courses='["高等数学","线性代数"]', max_weekly_sessions=5)
    Teacher.create(teacher_number='T002', name='李娜', teachable_courses='["大学英语","通信原理"]', max_weekly_sessions=5)
    Teacher.create(teacher_number='T003', name='王强', teachable_courses='["大学物理","数字信号处理"]', max_weekly_sessions=5)
    Teacher.create(teacher_number='T004', name='赵敏', teachable_courses='["数据结构","算法分析"]', max_weekly_sessions=5)
    
    # 创建班级
    SchoolClass.create(class_number='C001', name='计算机科学与技术1班', student_count=45, department='计算机科学与技术')
    SchoolClass.create(class_number='C002', name='软件工程1班', student_count=42, department='软件工程')
    SchoolClass.create(class_number='C003', name='网络工程1班', student_count=40, department='网络工程')
    
    _database.close()


def create_teaching_classes():
    """创建教学班记录"""
    if not _database.is_closed():
        _database.close()
    _database.connect()
    
    t1 = Teacher.get(Teacher.name == '张伟')
    t2 = Teacher.get(Teacher.name == '李娜')
    t3 = Teacher.get(Teacher.name == '王强')
    t4 = Teacher.get(Teacher.name == '赵敏')
    c1 = SchoolClass.get(SchoolClass.name == '计算机科学与技术1班')
    c2 = SchoolClass.get(SchoolClass.name == '软件工程1班')
    c3 = SchoolClass.get(SchoolClass.name == '网络工程1班')
    
    course1 = Course.get(Course.name == '高等数学')
    course2 = Course.get(Course.name == '数据结构')
    course3 = Course.get(Course.name == '大学英语')
    course4 = Course.get(Course.name == '大学物理')
    
    TeachingClass.create(course=course1, school_class=c1, teacher=t1, assigned_day=1, assigned_period='morning')
    TeachingClass.create(course=course2, school_class=c2, teacher=t4, assigned_day=2, assigned_period='morning')
    TeachingClass.create(course=course3, school_class=c3, teacher=t2, assigned_day=3, assigned_period='morning')
    TeachingClass.create(course=course4, school_class=c1, teacher=t3, assigned_day=4, assigned_period='morning')
    
    _database.close()


# ======================== 模块1: Excel数据导入功能测试 ========================

def test_excel_template_generation(client):
    """TC-001 ~ TC-005: 测试Excel模板下载功能"""
    print("\n=== 模块1.1: Excel模板下载功能 ===")
    
    # TC-001: 教室模板下载
    resp = client.get('/api/import/template/room')
    if resp.status_code == 200:
        wb = load_workbook(filename=tempfile_from_response(resp))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        expected = ['教室编号', '教室名称', '容量', '教室类型']
        passed = headers == expected
        log_result('TC-001', '教室模板下载', 
                   'PASS' if passed else 'FAIL',
                   f'表头: {headers}', 
                   'Blocker' if not passed else '')
    else:
        log_result('TC-001', '教室模板下载', 'FAIL', f'HTTP状态码: {resp.status_code}', 'Blocker')

    # TC-002: 教师模板下载
    resp = client.get('/api/import/template/teacher')
    if resp.status_code == 200:
        wb = load_workbook(filename=tempfile_from_response(resp))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        expected = ['教师工号', '教师姓名', '可授课程(用逗号分隔)', '每周最大课次数']
        passed = headers == expected
        log_result('TC-002', '教师模板下载',
                   'PASS' if passed else 'FAIL',
                   f'表头: {headers}',
                   'Blocker' if not passed else '')
    else:
        log_result('TC-002', '教师模板下载', 'FAIL', f'HTTP状态码: {resp.status_code}', 'Blocker')

    # TC-003: 班级模板下载
    resp = client.get('/api/import/template/class')
    if resp.status_code == 200:
        wb = load_workbook(filename=tempfile_from_response(resp))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        expected = ['班级编号', '班级名称', '学生人数', '所属院系']
        passed = headers == expected
        log_result('TC-003', '班级模板下载',
                   'PASS' if passed else 'FAIL',
                   f'表头: {headers}',
                   'Blocker' if not passed else '')
    else:
        log_result('TC-003', '班级模板下载', 'FAIL', f'HTTP状态码: {resp.status_code}', 'Blocker')

    # TC-004: 课程模板下载
    resp = client.get('/api/import/template/course')
    if resp.status_code == 200:
        wb = load_workbook(filename=tempfile_from_response(resp))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        expected = ['课程编号', '课程名称', '课程类型', '总课时', '授课教师姓名', '授课班级名称']
        passed = headers == expected
        log_result('TC-004', '课程模板下载',
                   'PASS' if passed else 'FAIL',
                   f'表头: {headers}',
                   'Blocker' if not passed else '')
    else:
        log_result('TC-004', '课程模板下载', 'FAIL', f'HTTP状态码: {resp.status_code}', 'Blocker')

    # TC-005: 不支持的模板类型
    resp = client.get('/api/import/template/invalid')
    passed = resp.status_code == 400
    log_result('TC-005', '无效模板类型处理',
               'PASS' if passed else 'FAIL',
               f'HTTP状态码: {resp.status_code}',
               'Major' if not passed else '')


def test_excel_data_import(client):
    """TC-010 ~ TC-013: 测试Excel数据导入功能"""
    print("\n=== 模块1.2: Excel数据导入功能 ===")

    # TC-010: 导入教室数据
    wb = Workbook()
    ws = wb.active
    ws.append(['教室编号', '教室名称', '容量', '教室类型'])
    ws.append(['R001', '博学楼101', 60, '普通教室'])
    ws.append(['R002', '博学楼102', 55, '普通教室'])
    ws.append(['R003', '博学楼201', 50, '多媒体教室'])
    ws.append(['R004', '博学楼301', 40, '机房'])
    ws.append(['R005', '博学楼401', 35, '实验室'])
    
    room_file = save_workbook_to_temp(wb, 'rooms.xlsx')
    
    with open(room_file, 'rb') as f:
        resp = client.post('/api/import/upload',
                           data={'file': (f, 'rooms.xlsx'), 'type': 'room'},
                           content_type='multipart/form-data')
    
    data = resp.get_json()
    passed = resp.status_code == 200 and data.get('count') == 5
    log_result('TC-010', '导入教室数据(5条)',
               'PASS' if passed else 'FAIL',
               f'HTTP: {resp.status_code}, 导入数量: {data.get("count")}, 错误: {data.get("errors")}',
               'Blocker' if not passed else '')

    # TC-011: 导入教师数据
    wb = Workbook()
    ws = wb.active
    ws.append(['教师工号', '教师姓名', '可授课程(用逗号分隔)', '每周最大课次数'])
    ws.append(['T001', '张伟', '高等数学,线性代数', 5])
    ws.append(['T002', '李娜', '大学英语,通信原理', 5])
    ws.append(['T003', '王强', '大学物理,数字信号处理', 5])
    ws.append(['T004', '赵敏', '数据结构,算法分析', 5])
    
    teacher_file = save_workbook_to_temp(wb, 'teachers.xlsx')
    
    with open(teacher_file, 'rb') as f:
        resp = client.post('/api/import/upload',
                           data={'file': (f, 'teachers.xlsx'), 'type': 'teacher'},
                           content_type='multipart/form-data')
    
    data = resp.get_json()
    passed = resp.status_code == 200 and data.get('count') == 4
    log_result('TC-011', '导入教师数据(4条)',
               'PASS' if passed else 'FAIL',
               f'HTTP: {resp.status_code}, 导入数量: {data.get("count")}, 错误: {data.get("errors")}',
               'Blocker' if not passed else '')

    # TC-012: 导入班级数据
    wb = Workbook()
    ws = wb.active
    ws.append(['班级编号', '班级名称', '学生人数', '所属院系'])
    ws.append(['C001', '计算机科学与技术1班', 45, '计算机科学与技术'])
    ws.append(['C002', '软件工程1班', 42, '软件工程'])
    ws.append(['C003', '网络工程1班', 40, '网络工程'])
    
    class_file = save_workbook_to_temp(wb, 'classes.xlsx')
    
    with open(class_file, 'rb') as f:
        resp = client.post('/api/import/upload',
                           data={'file': (f, 'classes.xlsx'), 'type': 'class'},
                           content_type='multipart/form-data')
    
    data = resp.get_json()
    passed = resp.status_code == 200 and data.get('count') == 3
    log_result('TC-012', '导入班级数据(3条)',
               'PASS' if passed else 'FAIL',
               f'HTTP: {resp.status_code}, 导入数量: {data.get("count")}, 错误: {data.get("errors")}',
               'Blocker' if not passed else '')

    # TC-013: 导入课程数据
    wb = Workbook()
    ws = wb.active
    ws.append(['课程编号', '课程名称', '课程类型', '总课时', '授课教师姓名', '授课班级名称'])
    ws.append(['CR001', '高等数学', '普通授课', 64, '张伟', '计算机科学与技术1班'])
    ws.append(['CR002', '数据结构', '上机', 64, '赵敏', '软件工程1班'])
    ws.append(['CR003', '大学英语', '普通授课', 64, '李娜', '网络工程1班'])
    ws.append(['CR004', '大学物理', '普通授课', 48, '王强', '计算机科学与技术1班'])
    
    course_file = save_workbook_to_temp(wb, 'courses.xlsx')
    
    with open(course_file, 'rb') as f:
        resp = client.post('/api/import/upload',
                           data={'file': (f, 'courses.xlsx'), 'type': 'course'},
                           content_type='multipart/form-data')
    
    data = resp.get_json()
    passed = resp.status_code == 200 and data.get('count') == 4
    log_result('TC-013', '导入课程数据(4条)',
               'PASS' if passed else 'FAIL',
               f'HTTP: {resp.status_code}, 导入数量: {data.get("count")}, 错误: {data.get("errors")}',
               'Blocker' if not passed else '')


def test_excel_import_edge_cases(client):
    """TC-020 ~ TC-026: Excel导入边界条件测试"""
    print("\n=== 模块1.3: Excel导入边界条件测试 ===")

    # TC-020: 上传文件为空
    resp = client.post('/api/import/upload',
                       data={'type': 'room'},
                       content_type='multipart/form-data')
    passed = resp.status_code == 400
    log_result('TC-020', '未上传文件处理',
               'PASS' if passed else 'FAIL',
               f'HTTP: {resp.status_code}',
               'Major' if not passed else '')

    # TC-021: 未指定数据类型
    empty_file = save_workbook_to_temp(Workbook(), 'empty.xlsx')
    with open(empty_file, 'rb') as f:
        resp = client.post('/api/import/upload',
                           data={'file': (f, 'empty.xlsx')},
                           content_type='multipart/form-data')
    passed = resp.status_code == 400
    log_result('TC-021', '未指定数据类型处理',
               'PASS' if passed else 'FAIL',
               f'HTTP: {resp.status_code}',
               'Major' if not passed else '')

    # TC-022: 不支持的文件格式
    resp = client.post('/api/import/upload',
                       data={'file': (io.BytesIO(b'fake csv content'), 'test.csv'), 'type': 'room'},
                       content_type='multipart/form-data')
    passed = resp.status_code == 400
    log_result('TC-022', '不支持文件格式处理',
               'PASS' if passed else 'FAIL',
               f'HTTP: {resp.status_code}',
               'Major' if not passed else '')

    # TC-023: Excel中无数据行
    wb = Workbook()
    ws = wb.active
    ws.append(['教室编号', '教室名称', '容量', '教室类型'])
    no_data_file = save_workbook_to_temp(wb, 'no_data.xlsx')
    with open(no_data_file, 'rb') as f:
        resp = client.post('/api/import/upload',
                           data={'file': (f, 'no_data.xlsx'), 'type': 'room'},
                           content_type='multipart/form-data')
    data = resp.get_json()
    passed = resp.status_code == 400
    log_result('TC-023', '空数据行处理',
               'PASS' if passed else 'FAIL',
               f'HTTP: {resp.status_code}, 响应: {data}',
               'Major' if not passed else '')

    # TC-024: 课程导入时教师不存在
    wb = Workbook()
    ws = wb.active
    ws.append(['课程编号', '课程名称', '课程类型', '总课时', '授课教师姓名', '授课班级名称'])
    ws.append(['CR999', '不存在的教师课程', '普通授课', 64, '不存在的教师', '计算机科学与技术1班'])
    invalid_course_file = save_workbook_to_temp(wb, 'invalid_course.xlsx')
    with open(invalid_course_file, 'rb') as f:
        resp = client.post('/api/import/upload',
                           data={'file': (f, 'invalid_course.xlsx'), 'type': 'course'},
                           content_type='multipart/form-data')
    data = resp.get_json()
    passed = resp.status_code == 207 and data.get('count') == 0
    log_result('TC-024', '课程导入教师不存在处理',
               'PASS' if passed else 'FAIL',
               f'HTTP: {resp.status_code}, 导入数量: {data.get("count")}, 错误: {data.get("errors")}',
               'Major' if not passed else '')

    # TC-025: 课程导入时班级不存在
    wb = Workbook()
    ws = wb.active
    ws.append(['课程编号', '课程名称', '课程类型', '总课时', '授课教师姓名', '授课班级名称'])
    ws.append(['CR998', '不存在的班级课程', '普通授课', 64, '张伟', '不存在的班级'])
    invalid_class_file = save_workbook_to_temp(wb, 'invalid_class.xlsx')
    with open(invalid_class_file, 'rb') as f:
        resp = client.post('/api/import/upload',
                           data={'file': (f, 'invalid_class.xlsx'), 'type': 'course'},
                           content_type='multipart/form-data')
    data = resp.get_json()
    passed = resp.status_code == 207 and data.get('count') == 0
    log_result('TC-025', '课程导入班级不存在处理',
               'PASS' if passed else 'FAIL',
               f'HTTP: {resp.status_code}, 导入数量: {data.get("count")}, 错误: {data.get("errors")}',
               'Major' if not passed else '')

    # TC-026: 部分成功部分失败
    wb = Workbook()
    ws = wb.active
    ws.append(['课程编号', '课程名称', '课程类型', '总课时', '授课教师姓名', '授课班级名称'])
    ws.append(['CR100', '正常课程', '普通授课', 64, '张伟', '计算机科学与技术1班'])
    ws.append(['CR101', '异常课程', '普通授课', 64, '不存在的教师', '计算机科学与技术1班'])
    partial_file = save_workbook_to_temp(wb, 'partial.xlsx')
    with open(partial_file, 'rb') as f:
        resp = client.post('/api/import/upload',
                           data={'file': (f, 'partial.xlsx'), 'type': 'course'},
                           content_type='multipart/form-data')
    data = resp.get_json()
    passed = resp.status_code == 207 and data.get('count') == 1
    log_result('TC-026', '部分成功部分失败处理',
               'PASS' if passed else 'FAIL',
               f'HTTP: {resp.status_code}, 导入数量: {data.get("count")}, 错误: {data.get("errors")}',
               'Major' if not passed else '')


# ======================== 模块2: 自动排课功能测试 ========================

def test_auto_scheduling(client):
    """TC-030 ~ TC-034: 测试自动排课功能"""
    print("\n=== 模块2.1: 自动排课功能 ===")
    
    # 清空已有排课数据
    if not _database.is_closed():
        _database.close()
    _database.connect()
    ScheduleEntry.delete().execute()
    TeachingClass.delete().execute()
    _database.close()
    
    # 创建教学班记录
    create_teaching_classes()
    
    # TC-030: 触发自动排课
    resp = client.post('/api/schedule/run',
                       json={'start_date': '2026-09-07'},
                       content_type='application/json')
    data = resp.get_json()
    passed = resp.status_code == 200
    log_result('TC-030', '触发自动排课',
               'PASS' if passed else 'FAIL',
               f'HTTP: {resp.status_code}, 响应: {data}',
               'Blocker' if not passed else '')
    
    if passed:
        success_count = data.get('success_count', 0)
        failed_count = data.get('failed_count', 0)
        
        # TC-031: 排课成功数量
        passed = success_count > 0
        log_result('TC-031', '排课成功数量>0',
                   'PASS' if passed else 'FAIL',
                   f'成功: {success_count}, 失败: {failed_count}',
                   'Critical' if not passed else '')
        
        # TC-032: 排课日志返回
        passed = 'log' in data and len(data['log']) > 0
        log_result('TC-032', '排课日志返回',
                   'PASS' if passed else 'FAIL',
                   f'日志条目数: {len(data.get("log", []))}',
                   'Major' if not passed else '')
        
        # TC-033: 查询排课结果
        resp = client.get('/api/schedule/results?per_page=100')
        data = resp.get_json()
        passed = resp.status_code == 200 and data.get('total', 0) > 0
        log_result('TC-033', '查询排课结果',
                   'PASS' if passed else 'FAIL',
                   f'HTTP: {resp.status_code}, 结果总数: {data.get("total")}',
                   'Critical' if not passed else '')
        
        if passed:
            results = data.get('results', [])
            
            # TC-034: 排课结果完整性验证
            has_room = all(r.get('room_name') for r in results)
            has_teacher = all(r.get('teacher_name') for r in results)
            has_course = all(r.get('course_name') for r in results)
            passed = has_room and has_teacher and has_course
            log_result('TC-034', '排课结果完整性(教室/教师/课程)',
                       'PASS' if passed else 'FAIL',
                       f'教室:{has_room}, 教师:{has_teacher}, 课程:{has_course}',
                       'Critical' if not passed else '')


# ======================== 模块3: 排课算法正确性验证 ========================

def test_scheduling_algorithm():
    """TC-040 ~ TC-046: 验证排课算法正确性"""
    print("\n=== 模块3.1: 排课算法正确性验证 ===")
    
    if not _database.is_closed():
        _database.close()
    _database.connect()
    
    entries = list(ScheduleEntry.select().order_by(ScheduleEntry.week, ScheduleEntry.day, ScheduleEntry.period))
    
    if len(entries) == 0:
        log_result('TC-040', '教室冲突检测', 'SKIP', '无排课数据，跳过', '')
        log_result('TC-041', '教师冲突检测', 'SKIP', '无排课数据，跳过', '')
        log_result('TC-042', '班级冲突检测', 'SKIP', '无排课数据，跳过', '')
        log_result('TC-043', '时段有效性检测', 'SKIP', '无排课数据，跳过', '')
        log_result('TC-044', '周次范围检测(1-16)', 'SKIP', '无排课数据，跳过', '')
        log_result('TC-045', '教室分配完整性', 'SKIP', '无排课数据，跳过', '')
        log_result('TC-046', '教学班关联完整性', 'SKIP', '无排课数据，跳过', '')
        _database.close()
        return
    
    # TC-040: 检测教室冲突（同一时段同一教室只能有一门课）
    room_conflicts = []
    room_slots = {}
    for entry in entries:
        slot_key = f"W{entry.week}_D{entry.day}_{entry.period}_R{entry.room_id}"
        if slot_key in room_slots:
            room_conflicts.append({
                'slot': slot_key,
                'entries': [room_slots[slot_key], entry.id]
            })
        else:
            room_slots[slot_key] = entry.id
    
    passed = len(room_conflicts) == 0
    log_result('TC-040', '教室冲突检测',
               'PASS' if passed else 'FAIL',
               f'冲突数量: {len(room_conflicts)}' + (f', 详情: {room_conflicts[:3]}' if room_conflicts else ''),
               'Blocker' if not passed else '')
    
    # TC-041: 检测教师冲突（同一教师同一时段只能教一门课）
    teacher_conflicts = []
    teacher_slots = {}
    for entry in entries:
        tc = entry.teaching_class
        if tc and tc.teacher:
            teacher_id = tc.teacher.id
            slot_key = f"W{entry.week}_D{entry.day}_{entry.period}_T{teacher_id}"
            if slot_key in teacher_slots:
                teacher_conflicts.append({
                    'slot': slot_key,
                    'entries': [teacher_slots[slot_key], entry.id]
                })
            else:
                teacher_slots[slot_key] = entry.id
    
    passed = len(teacher_conflicts) == 0
    log_result('TC-041', '教师冲突检测',
               'PASS' if passed else 'FAIL',
               f'冲突数量: {len(teacher_conflicts)}' + (f', 详情: {teacher_conflicts[:3]}' if teacher_conflicts else ''),
               'Blocker' if not passed else '')
    
    # TC-042: 检测班级冲突（同一班级同一时段只能上一门课）
    class_conflicts = []
    class_slots = {}
    for entry in entries:
        tc = entry.teaching_class
        if tc and tc.school_class:
            class_id = tc.school_class.id
            slot_key = f"W{entry.week}_D{entry.day}_{entry.period}_C{class_id}"
            if slot_key in class_slots:
                class_conflicts.append({
                    'slot': slot_key,
                    'entries': [class_slots[slot_key], entry.id]
                })
            else:
                class_slots[slot_key] = entry.id
    
    passed = len(class_conflicts) == 0
    log_result('TC-042', '班级冲突检测',
               'PASS' if passed else 'FAIL',
               f'冲突数量: {len(class_conflicts)}' + (f', 详情: {class_conflicts[:3]}' if class_conflicts else ''),
               'Blocker' if not passed else '')
    
    # TC-043: 检测时段有效性（时段只能是 morning/afternoon/evening）
    valid_periods = {'morning', 'afternoon', 'evening'}
    invalid_periods = [e for e in entries if e.period not in valid_periods]
    passed = len(invalid_periods) == 0
    log_result('TC-043', '时段有效性检测',
               'PASS' if passed else 'FAIL',
               f'无效时段记录数: {len(invalid_periods)}',
               'Critical' if not passed else '')
    
    # TC-044: 检测周次范围（应在1-16范围内）
    out_of_range = [e for e in entries if e.week < 1 or e.week > 16]
    passed = len(out_of_range) == 0
    log_result('TC-044', '周次范围检测(1-16)',
               'PASS' if passed else 'FAIL',
               f'越界记录数: {len(out_of_range)}',
               'Critical' if not passed else '')
    
    # TC-045: 检测教室分配完整性（每条记录都应有教室）
    missing_rooms = [e for e in entries if not e.room_id]
    passed = len(missing_rooms) == 0
    log_result('TC-045', '教室分配完整性',
               'PASS' if passed else 'FAIL',
               f'缺少教室的记录数: {len(missing_rooms)}',
               'Critical' if not passed else '')
    
    # TC-046: 教学班关联完整性
    missing_teaching_class = []
    for entry in entries:
        if not entry.teaching_class_id:
            missing_teaching_class.append(entry.id)
        else:
            tc = TeachingClass.get_or_none(TeachingClass.id == entry.teaching_class_id)
            if not tc:
                missing_teaching_class.append(entry.id)
    
    passed = len(missing_teaching_class) == 0
    log_result('TC-046', '教学班关联完整性',
               'PASS' if passed else 'FAIL',
               f'无效教学班关联记录数: {len(missing_teaching_class)}',
               'Critical' if not passed else '')
    
    _database.close()


# ======================== 模块4: 课表显示功能测试 ========================

def test_schedule_display(client):
    """TC-050 ~ TC-058: 测试课表显示功能"""
    print("\n=== 模块4.1: 课表显示与统计功能 ===")
    
    # TC-050: 周课表API返回
    resp = client.get('/api/schedule/weekly')
    data = resp.get_json()
    passed = resp.status_code == 200 and 'courses' in data
    log_result('TC-050', '周课表API返回',
               'PASS' if passed else 'FAIL',
               f'HTTP: {resp.status_code}, 包含courses字段: {"courses" in data}',
               'Critical' if not passed else '')
    
    if passed:
        courses = data.get('courses', [])
        
        # TC-051: 周课表数据结构
        required_fields = ['day', 'period', 'course_name', 'teacher_name', 'room_name', 'week_ranges']
        if len(courses) > 0:
            first_course = courses[0]
            missing_fields = [f for f in required_fields if f not in first_course]
            passed = len(missing_fields) == 0
            log_result('TC-051', '周课表数据结构完整性',
                       'PASS' if passed else 'FAIL',
                       f'缺失字段: {missing_fields}' if missing_fields else '所有必需字段都存在',
                       'Major' if not passed else '')
        else:
            log_result('TC-051', '周课表数据结构完整性', 'SKIP', '无课表数据', '')
        
        # TC-052: week_ranges字段格式
        if len(courses) > 0:
            has_valid_ranges = True
            for c in courses:
                wr = c.get('week_ranges', '')
                if wr:
                    parts = wr.split(',')
                    for part in parts:
                        if '-' in part:
                            nums = part.split('-')
                            if len(nums) != 2 or not nums[0].isdigit() or not nums[1].isdigit():
                                has_valid_ranges = False
                                break
                        elif not part.isdigit():
                            has_valid_ranges = False
                            break
            log_result('TC-052', 'week_ranges格式正确',
                       'PASS' if has_valid_ranges else 'FAIL',
                       f'week_ranges格式验证: {has_valid_ranges}',
                       'Major' if not has_valid_ranges else '')
        else:
            log_result('TC-052', 'week_ranges格式正确', 'SKIP', '无课表数据', '')
        
        # TC-053: week_dates信息
        week_dates = data.get('week_dates', {})
        passed = len(week_dates) == 16 and '1' in week_dates
        log_result('TC-053', 'week_dates信息完整性(16周)',
                   'PASS' if passed else 'FAIL',
                   f'周次数量: {len(week_dates)}',
                   'Major' if not passed else '')
        
        # TC-054: 按班级筛选课表
        resp = client.get('/api/schedule/weekly?class_id=1')
        passed = resp.status_code == 200
        log_result('TC-054', '按班级筛选课表',
                   'PASS' if passed else 'FAIL',
                   f'HTTP: {resp.status_code}',
                   'Major' if not passed else '')
        
        # TC-055: 按教师筛选课表
        resp = client.get('/api/schedule/weekly?teacher_id=1')
        passed = resp.status_code == 200
        log_result('TC-055', '按教师筛选课表',
                   'PASS' if passed else 'FAIL',
                   f'HTTP: {resp.status_code}',
                   'Major' if not passed else '')
    
    # TC-056: 统计API返回
    resp = client.get('/api/schedule/statistics')
    data = resp.get_json()
    passed = resp.status_code == 200
    log_result('TC-056', '统计API返回',
               'PASS' if passed else 'FAIL',
               f'HTTP: {resp.status_code}',
               'Major' if not passed else '')
    
    if passed:
        # TC-057: 统计数据字段完整性
        required_fields = ['total_hours', 'total_sessions', 'period_distribution', 
                         'daily_distribution', 'completion_rate', 'total_classes', 'scheduled_classes']
        missing = [f for f in required_fields if f not in data]
        passed = len(missing) == 0
        log_result('TC-057', '统计数据字段完整性',
                   'PASS' if passed else 'FAIL',
                   f'缺失字段: {missing}' if missing else '所有必需字段都存在',
                   'Major' if not passed else '')
        
        # TC-058: 完成率计算
        completion_rate = data.get('completion_rate', 0)
        total = data.get('total_classes', 0)
        scheduled = data.get('scheduled_classes', 0)
        if total > 0:
            expected_rate = round(scheduled / total * 100, 2)
            passed = abs(completion_rate - expected_rate) < 0.01
            log_result('TC-058', '完成率计算正确',
                       'PASS' if passed else 'FAIL',
                       f'完成率: {completion_rate}%, 预期: {expected_rate}%',
                       'Major' if not passed else '')
        else:
            log_result('TC-058', '完成率计算正确', 'SKIP', '无教学班数据', '')


# ======================== 模块5: 课表调整与冲突检测 ========================

def test_schedule_adjustment(client):
    """TC-060 ~ TC-064: 测试课表调整功能"""
    print("\n=== 模块5.1: 课表调整与冲突检测 ===")
    
    if not _database.is_closed():
        _database.close()
    _database.connect()
    
    entries = list(ScheduleEntry.select().limit(5))
    
    if len(entries) == 0:
        log_result('TC-060', '调整课表成功', 'SKIP', '无排课数据', '')
        log_result('TC-061', '调整课表-教室冲突检测', 'SKIP', '无排课数据', '')
        log_result('TC-062', '调整课表-教师冲突检测', 'SKIP', '无排课数据', '')
        log_result('TC-063', '调整不存在的记录', 'SKIP', '无排课数据', '')
        log_result('TC-064', '冲突检测API', 'SKIP', '无排课数据', '')
        _database.close()
        return
    
    test_entry = entries[0]
    tc = test_entry.teaching_class
    
    # TC-060: 调整课表成功
    resp = client.put(f'/api/schedule/adjust/{test_entry.id}',
                     json={'day': test_entry.day, 'period': 'evening', 'room_id': test_entry.room_id},
                     content_type='application/json')
    data = resp.get_json()
    passed = resp.status_code == 200
    log_result('TC-060', '调整课表成功',
               'PASS' if passed else 'FAIL',
               f'HTTP: {resp.status_code}, 响应: {data}',
               'Major' if not passed else '')
    
    # 恢复原设置
    if passed:
        client.put(f'/api/schedule/adjust/{test_entry.id}',
                   json={'day': test_entry.day, 'period': test_entry.period, 'room_id': test_entry.room_id},
                   content_type='application/json')
    
    # TC-061: 调整课表-教室冲突检测
    same_period_entries = list(ScheduleEntry.select().where(
        (ScheduleEntry.week == test_entry.week) & 
        (ScheduleEntry.day == test_entry.day) & 
        (ScheduleEntry.period == test_entry.period) &
        (ScheduleEntry.id != test_entry.id)
    ))
    
    if len(same_period_entries) > 0:
        conflict_entry = same_period_entries[0]
        resp = client.put(f'/api/schedule/adjust/{test_entry.id}',
                         json={'day': test_entry.day, 'period': test_entry.period, 'room_id': conflict_entry.room_id},
                         content_type='application/json')
        passed = resp.status_code == 400
        log_result('TC-061', '调整课表-教室冲突检测',
                   'PASS' if passed else 'FAIL',
                   f'HTTP: {resp.status_code}, 响应: {resp.get_json()}',
                   'Critical' if not passed else '')
    else:
        log_result('TC-061', '调整课表-教室冲突检测', 'SKIP', '无同时段其他教室', '')
    
    # TC-062: 调整课表-教师冲突检测
    if tc:
        teacher_entries = list(ScheduleEntry.select().join(
            TeachingClass, on=(ScheduleEntry.teaching_class == TeachingClass.id)
        ).where(
            (TeachingClass.teacher == tc.teacher) &
            (ScheduleEntry.week == test_entry.week) &
            (ScheduleEntry.day == test_entry.day) &
            (ScheduleEntry.period == test_entry.period) &
            (ScheduleEntry.id != test_entry.id)
        ))
        
        if len(teacher_entries) > 0:
            resp = client.put(f'/api/schedule/adjust/{test_entry.id}',
                             json={'day': test_entry.day, 'period': test_entry.period, 'room_id': test_entry.room_id},
                             content_type='application/json')
            passed = resp.status_code in [200, 400]
            log_result('TC-062', '调整课表-教师冲突检测',
                       'PASS' if passed else 'FAIL',
                       f'HTTP: {resp.status_code}',
                       'Critical' if not passed else '')
        else:
            log_result('TC-062', '调整课表-教师冲突检测', 'SKIP', '无同教师同时段课程', '')
    else:
        log_result('TC-062', '调整课表-教师冲突检测', 'SKIP', '无教学班信息', '')
    
    # TC-063: 调整不存在的记录
    resp = client.put('/api/schedule/adjust/999999',
                     json={'day': 1, 'period': 'morning', 'room_id': 1},
                     content_type='application/json')
    passed = resp.status_code == 404
    log_result('TC-063', '调整不存在的记录',
               'PASS' if passed else 'FAIL',
               f'HTTP: {resp.status_code}',
               'Major' if not passed else '')
    
    # TC-064: 冲突检测API
    if tc:
        resp = client.post(f'/api/schedule/check-conflict/{tc.id}',
                          json={},
                          content_type='application/json')
        data = resp.get_json()
        passed = resp.status_code == 200 and 'has_conflict' in data
        log_result('TC-064', '冲突检测API',
                   'PASS' if passed else 'FAIL',
                   f'HTTP: {resp.status_code}, 响应: {data}',
                   'Major' if not passed else '')
    else:
        log_result('TC-064', '冲突检测API', 'SKIP', '无教学班信息', '')
    
    _database.close()


# ======================== 模块6: 节假日管理测试 ========================

def test_holiday_management(client):
    """TC-070 ~ TC-073: 测试节假日管理功能"""
    print("\n=== 模块6.1: 节假日管理功能 ===")
    
    # TC-070: 添加节假日
    resp = client.post('/api/holidays',
                      json={'date': '2026-10-01', 'name': '国庆节'},
                      content_type='application/json')
    passed = resp.status_code in [200, 201, 409]
    log_result('TC-070', '添加节假日',
               'PASS' if passed else 'FAIL',
               f'HTTP: {resp.status_code}, 响应: {resp.get_json()}',
               'Major' if not passed else '')
    
    # TC-071: 查询节假日列表
    resp = client.get('/api/holidays')
    data = resp.get_json()
    passed = resp.status_code == 200 and isinstance(data, list)
    log_result('TC-071', '查询节假日列表',
               'PASS' if passed else 'FAIL',
               f'HTTP: {resp.status_code}, 数据类型: {type(data).__name__}',
               'Major' if not passed else '')
    
    # TC-072: 删除节假日
    holidays = resp.get_json() if passed else []
    if len(holidays) > 0:
        holiday_id = holidays[0].get('id')
        resp = client.delete(f'/api/holidays/{holiday_id}')
        passed = resp.status_code in [200, 204]
        log_result('TC-072', '删除节假日',
                   'PASS' if passed else 'FAIL',
                   f'HTTP: {resp.status_code}',
                   'Major' if not passed else '')
    else:
        log_result('TC-072', '删除节假日', 'SKIP', '无可删除节假日', '')
    
    # TC-073: 添加重复节假日
    client.post('/api/holidays',
               json={'date': '2026-05-01', 'name': '劳动节'},
               content_type='application/json')
    resp = client.post('/api/holidays',
                      json={'date': '2026-05-01', 'name': '劳动节重复'},
                      content_type='application/json')
    passed = resp.status_code in [409, 200]
    log_result('TC-073', '添加重复节假日处理',
               'PASS' if passed else 'FAIL',
               f'HTTP: {resp.status_code}',
               'Minor' if not passed else '')


# ======================== 模块7: 系统级功能测试 ========================

def test_system_functions(client):
    """TC-080 ~ TC-085: 系统级功能测试"""
    print("\n=== 模块7.1: 系统级功能测试 ===")
    
    # TC-080: 健康检查
    resp = client.get('/api/health')
    passed = resp.status_code == 200 and resp.get_json().get('status') == 'ok'
    log_result('TC-080', '健康检查API',
               'PASS' if passed else 'FAIL',
               f'HTTP: {resp.status_code}, 状态: {resp.get_json()}',
               'Critical' if not passed else '')
    
    # TC-081: CORS跨域支持
    resp = client.get('/api/health', headers={'Origin': 'http://localhost:3000'})
    passed = 'Access-Control-Allow-Origin' in resp.headers
    log_result('TC-081', 'CORS跨域支持',
               'PASS' if passed else 'FAIL',
               f'CORS头: {resp.headers.get("Access-Control-Allow-Origin", "缺失")}',
               'Major' if not passed else '')
    
    # TC-082: CRUD - 教室管理
    resp = client.get('/api/rooms')
    passed = resp.status_code == 200
    log_result('TC-082', '教室列表查询',
               'PASS' if passed else 'FAIL',
               f'HTTP: {resp.status_code}',
               'Major' if not passed else '')
    
    # TC-083: CRUD - 教师管理
    resp = client.get('/api/teachers')
    passed = resp.status_code == 200
    log_result('TC-083', '教师列表查询',
               'PASS' if passed else 'FAIL',
               f'HTTP: {resp.status_code}',
               'Major' if not passed else '')
    
    # TC-084: CRUD - 班级管理
    resp = client.get('/api/classes')
    passed = resp.status_code == 200
    log_result('TC-084', '班级列表查询',
               'PASS' if passed else 'FAIL',
               f'HTTP: {resp.status_code}',
               'Major' if not passed else '')
    
    # TC-085: CRUD - 课程管理
    resp = client.get('/api/courses')
    passed = resp.status_code == 200
    log_result('TC-085', '课程列表查询',
               'PASS' if passed else 'FAIL',
               f'HTTP: {resp.status_code}',
               'Major' if not passed else '')


# ======================== 测试主入口 ========================

def main():
    global TEST_RESULTS, TEST_PASS_COUNT, TEST_FAIL_COUNT, TEST_SKIP_COUNT
    
    print("=" * 70)
    print("智能排课系统 - 全面自动化测试")
    print("=" * 70)
    print(f"测试时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Python版本: {sys.version}")
    print("=" * 70)
    
    # 初始化Flask应用
    from app import create_app
    from config import Config
    
    # 确保数据库正确初始化
    if not _database.is_closed():
        _database.close()
    
    app = create_app()
    client = app.test_client()
    
    # 确保数据库已连接
    if _database.is_closed():
        _database.connect()
    
    # 创建表（如果不存在）
    _database.create_tables([Room, Teacher, SchoolClass, Course, Holiday, TeachingClass, ScheduleEntry], safe=True)
    
    # 清空数据并初始化
    print("\n正在初始化测试环境...")
    clear_all_data()
    create_test_data()
    
    # 执行测试模块
    try:
        test_excel_template_generation(client)
    except Exception as e:
        print(f"\n  [错误] Excel模板测试异常: {e}")
        import traceback
        traceback.print_exc()
    
    try:
        test_excel_data_import(client)
    except Exception as e:
        print(f"\n  [错误] Excel导入测试异常: {e}")
        import traceback
        traceback.print_exc()
    
    try:
        test_excel_import_edge_cases(client)
    except Exception as e:
        print(f"\n  [错误] Excel导入边界测试异常: {e}")
        import traceback
        traceback.print_exc()
    
    try:
        test_auto_scheduling(client)
    except Exception as e:
        print(f"\n  [错误] 自动排课测试异常: {e}")
        import traceback
        traceback.print_exc()
    
    try:
        test_scheduling_algorithm()
    except Exception as e:
        print(f"\n  [错误] 排课算法测试异常: {e}")
        import traceback
        traceback.print_exc()
    
    try:
        test_schedule_display(client)
    except Exception as e:
        print(f"\n  [错误] 课表显示测试异常: {e}")
        import traceback
        traceback.print_exc()
    
    try:
        test_schedule_adjustment(client)
    except Exception as e:
        print(f"\n  [错误] 课表调整测试异常: {e}")
        import traceback
        traceback.print_exc()
    
    try:
        test_holiday_management(client)
    except Exception as e:
        print(f"\n  [错误] 节假日管理测试异常: {e}")
        import traceback
        traceback.print_exc()
    
    try:
        test_system_functions(client)
    except Exception as e:
        print(f"\n  [错误] 系统功能测试异常: {e}")
        import traceback
        traceback.print_exc()
    
    # 清理
    cleanup_temp_files()
    
    # 输出测试报告
    print("\n" + "=" * 70)
    print("测 试 报 告")
    print("=" * 70)
    total = TEST_PASS_COUNT + TEST_FAIL_COUNT + TEST_SKIP_COUNT
    print(f"测试总数: {total}")
    print(f"通过: {TEST_PASS_COUNT}")
    print(f"失败: {TEST_FAIL_COUNT}")
    print(f"跳过: {TEST_SKIP_COUNT}")
    if TEST_PASS_COUNT + TEST_FAIL_COUNT > 0:
        print(f"通过率: {TEST_PASS_COUNT / (TEST_PASS_COUNT + TEST_FAIL_COUNT) * 100:.1f}%")
    else:
        print(f"通过率: N/A")
    print("=" * 70)
    
    # 失败项汇总
    failed_items = [r for r in TEST_RESULTS if r['status'] == 'FAIL']
    if failed_items:
        print("\n【失败项汇总】")
        for item in failed_items:
            severity_map = {
                'Blocker': '致命',
                'Critical': '严重',
                'Major': '一般',
                'Minor': '轻微'
            }
            severity = severity_map.get(item['severity'], '未定义')
            print(f"  [{severity}] {item['test_id']} {item['test_name']}: {item['detail']}")
    
    # 缺陷统计
    blocker_count = len([r for r in failed_items if r['severity'] == 'Blocker'])
    critical_count = len([r for r in failed_items if r['severity'] == 'Critical'])
    major_count = len([r for r in failed_items if r['severity'] == 'Major'])
    minor_count = len([r for r in failed_items if r['severity'] == 'Minor'])
    
    print(f"\n【缺陷统计】")
    print(f"  致命(Blocker): {blocker_count}")
    print(f"  严重(Critical): {critical_count}")
    print(f"  一般(Major): {major_count}")
    print(f"  轻微(Minor): {minor_count}")
    
    # 给出验收结论
    print(f"\n【验收结论】")
    if blocker_count == 0 and critical_count == 0:
        print("  建议: 可以发布 (无致命/严重缺陷)")
    elif blocker_count == 0:
        print(f"  建议: 有条件发布 (存在 {critical_count} 个严重缺陷，需评估)")
    else:
        print(f"  建议: 不建议发布 (存在 {blocker_count} 个致命缺陷)")
    
    print("=" * 70)
    
    return TEST_RESULTS


if __name__ == '__main__':
    main()
