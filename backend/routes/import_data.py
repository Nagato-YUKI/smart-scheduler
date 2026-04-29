from flask import Blueprint, request, jsonify, send_file
import io
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from peewee_manager import Room, Teacher, SchoolClass, Course, TeachingClass, Holiday, _database

import_bp = Blueprint('import', __name__, url_prefix='/api/import')

# 定义模板配置
TEMPLATE_CONFIG = {
    'room': {
        'filename': '教室_导入模板.xlsx',
        'headers': ['教室编号', '教室名称', '容量', '教室类型'],
        'fields': ['room_number', 'name', 'capacity', 'room_type'],
        'examples': [
            ['R001', '博学楼101', 60, '普通教室'],
            ['R002', '博学楼102', 55, '普通教室'],
            ['R003', '博学楼201', 50, '多媒体教室'],
            ['R004', '博学楼301', 40, '机房'],
            ['R005', '博学楼401', 35, '实验室'],
        ],
        'tips': '教室类型可选：普通教室、多媒体教室、机房、实验室'
    },
    'teacher': {
        'filename': '教师_导入模板.xlsx',
        'headers': ['教师工号', '教师姓名', '可授课程(用逗号分隔)', '每周最大课次数'],
        'fields': ['teacher_number', 'name', 'teachable_courses', 'max_weekly_sessions'],
        'examples': [
            ['T001', '张伟', '高等数学,线性代数', 5],
            ['T002', '李娜', '大学英语,通信原理', 5],
            ['T003', '王强', '大学物理,数字信号处理', 5],
        ],
        'tips': '可授课程用逗号分隔，如：高等数学,线性代数,概率论'
    },
    'class': {
        'filename': '班级_导入模板.xlsx',
        'headers': ['班级编号', '班级名称', '学生人数', '所属院系'],
        'fields': ['class_number', 'name', 'student_count', 'department'],
        'examples': [
            ['C001', '计算机科学与技术1班', 45, '计算机科学与技术'],
            ['C002', '软件工程1班', 42, '软件工程'],
            ['C003', '网络工程1班', 40, '网络工程'],
        ],
        'tips': '院系名称需与系统中已有的院系一致'
    },
    'course': {
        'filename': '课程_导入模板.xlsx',
        'headers': ['课程编号', '课程名称', '课程类型', '总课时', '授课教师姓名', '授课班级名称'],
        'fields': ['course_number', 'name', 'course_type', 'total_hours', 'teacher_name', 'class_name'],
        'examples': [
            ['CR001', '高等数学', '普通授课', 64, '张伟', '计算机科学与技术1班'],
            ['CR002', '数据结构', '上机', 64, '陈明', '软件工程1班'],
            ['CR003', '大学英语', '普通授课', 64, '李娜', '网络工程1班'],
        ],
        'tips': '课程类型可选：普通授课、上机、实验；教师姓名和班级名称需与系统中已有数据一致'
    },
    'holiday': {
        'filename': '节假日_导入模板.xlsx',
        'headers': ['日期', '节假日名称', '备注'],
        'fields': ['date', 'name', 'remark'],
        'examples': [
            ['2026-09-27', '中秋节', '中秋节放假'],
            ['2026-10-01', '国庆节', '国庆节放假'],
            ['2027-01-01', '元旦', '元旦放假'],
        ],
        'tips': '日期格式为YYYY-MM-DD'
    },
    'comprehensive': {
        'filename': '综合导入模板.xlsx',
        'is_comprehensive': True,
    }
}


def create_excel_template(type_name):
    """创建Excel模板"""
    config = TEMPLATE_CONFIG.get(type_name)
    if not config:
        return None, None, '不支持的数据类型'

    if config.get('is_comprehensive'):
        return create_comprehensive_template()

    wb = Workbook()
    ws = wb.active
    ws.title = '导入模板'

    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    example_font = Font(name='微软雅黑', size=10, color='666666')
    tips_font = Font(name='微软雅黑', size=9, italic=True, color='FF6600')

    for col, header in enumerate(config['headers'], 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, example in enumerate(config['examples'], 2):
        for col_idx, value in enumerate(example, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = example_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    for col in range(1, len(config['headers']) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    row_tips = len(config['examples']) + 3
    ws.cell(row=row_tips, column=1, value='使用说明：')
    ws.cell(row=row_tips, column=1).font = Font(name='微软雅黑', size=10, bold=True)
    ws.cell(row=row_tips + 1, column=1, value=f'1. {config["tips"]}')
    ws.cell(row=row_tips + 1, column=1).font = tips_font
    ws.cell(row=row_tips + 2, column=1, value='2. 请填写实际数据，删除示例行')
    ws.cell(row=row_tips + 2, column=1).font = tips_font
    ws.cell(row=row_tips + 3, column=1, value='3. 导入时将覆盖已有数据，请谨慎操作')
    ws.cell(row=row_tips + 3, column=1).font = tips_font

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output, config['filename'], None


def create_comprehensive_template():
    """创建综合导入模板（包含所有工作表）"""
    wb = Workbook()

    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    example_font = Font(name='微软雅黑', size=10, color='666666')
    tips_font = Font(name='微软雅黑', size=9, italic=True, color='FF6600')

    sheet_configs = [
        ('教室数据', ['教室编号', '教室名称', '容量', '教室类型'], [
            ['R001', '博学楼101', 60, '普通教室'],
            ['R002', '博学楼102', 55, '普通教室'],
            ['R003', '博学楼201', 50, '多媒体教室'],
        ]),
        ('教师数据', ['教师工号', '教师姓名', '可授课程(用逗号分隔)', '每周最大课次数'], [
            ['T001', '张伟', '高等数学,线性代数', 5],
            ['T002', '李娜', '大学英语,通信原理', 5],
        ]),
        ('班级数据', ['班级编号', '班级名称', '学生人数', '所属院系'], [
            ['C001', '计算机科学与技术1班', 45, '计算机科学与技术'],
            ['C002', '软件工程1班', 42, '软件工程'],
        ]),
        ('节假日数据', ['日期', '节假日名称', '备注'], [
            ['2026-09-27', '中秋节', '中秋节放假'],
            ['2026-10-01', '国庆节', '国庆节放假'],
        ]),
        ('课程数据', ['课程编号', '课程名称', '课程类型', '总课时', '授课教师姓名', '授课班级名称'], [
            ['CR001', '高等数学', '普通授课', 64, '张伟', '计算机科学与技术1班'],
            ['CR002', '数据结构', '上机', 64, '陈明', '软件工程1班'],
        ]),
    ]

    for idx, (sheet_name, headers, examples) in enumerate(sheet_configs):
        if idx == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_idx, example in enumerate(examples, 2):
            for col_idx, value in enumerate(example, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = example_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output, '综合导入模板.xlsx', None


@import_bp.route('/template/<type_name>', methods=['GET'])
def download_template(type_name):
    """下载Excel导入模板"""
    output, filename, error = create_excel_template(type_name)
    if error:
        return jsonify({'error': error}), 400
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


@import_bp.route('/upload', methods=['POST'])
def upload_data():
    """上传并导入Excel数据"""
    if 'file' not in request.files:
        return jsonify({'error': '未找到上传文件'}), 400

    file = request.files['file']
    data_type = request.form.get('type')

    if not data_type:
        return jsonify({'error': '未指定数据类型'}), 400

    if file.filename == '':
        return jsonify({'error': '文件名为空'}), 400

    allowed_extensions = {'.xlsx', '.xls'}
    file_ext = file.filename[file.filename.rfind('.'):].lower()
    if file_ext not in allowed_extensions:
        return jsonify({'error': '仅支持 .xlsx 或 .xls 格式的Excel文件'}), 400

    try:
        from openpyxl import load_workbook
        wb = load_workbook(file)
        
        # 如果是综合导入（comprehensive），处理所有工作表
        if data_type == 'comprehensive':
            return import_comprehensive(wb)
        
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        
        data_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(cell is None for cell in row):
                continue
            data_rows.append(list(row))

        if not data_rows:
            return jsonify({'error': '文件中没有数据行'}), 400

        imported_count = 0
        errors = []

        if data_type == 'room':
            imported, updated, skipped, errors = import_rooms(data_rows, headers)
            imported_count = imported + updated
        elif data_type == 'teacher':
            imported, updated, skipped, errors = import_teachers(data_rows, headers)
            imported_count = imported + updated
        elif data_type == 'class':
            imported, updated, skipped, errors = import_classes(data_rows, headers)
            imported_count = imported + updated
        elif data_type == 'course':
            imported, updated, skipped, errors = import_courses(data_rows, headers)
            imported_count = imported + updated
        elif data_type == 'holiday':
            imported, errors = import_holidays(data_rows, headers)
            imported_count = imported
        else:
            return jsonify({'error': f'不支持的数据类型: {data_type}'}), 400

        if errors:
            return jsonify({
                'count': imported_count,
                'errors': '; '.join(errors[:5])
            }), 207

        return jsonify({
            'count': imported_count,
            'errors': None
        })

    except Exception as e:
        return jsonify({'error': f'导入失败: {str(e)}'}), 500


def import_holidays(data_rows, headers):
    """导入节假日数据"""
    imported = 0
    errors = []

    def find_header(key):
        for h in headers:
            if h and key in str(h):
                return headers.index(h)
        return -1

    date_idx = find_header('日期')
    name_idx = find_header('节假日名称') if find_header('节假日名称') >= 0 else 1
    remark_idx = find_header('备注') if find_header('备注') >= 0 else -1

    for i, row in enumerate(data_rows, start=2):
        try:
            if len(row) < 2 or row[0] is None:
                continue

            date_str = str(row[date_idx]).strip() if date_idx >= 0 else ''
            name = str(row[name_idx]).strip() if name_idx >= 0 else ''
            remark = str(row[remark_idx]).strip() if remark_idx >= 0 and remark_idx < len(row) and row[remark_idx] else ''

            if not date_str or not name:
                errors.append(f'第{i}行：日期或节假日名称为空')
                continue

            # 解析日期
            try:
                from datetime import datetime
                date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                errors.append(f'第{i}行：日期格式错误，应为 YYYY-MM-DD')
                continue

            # 检查是否已存在
            existing = Holiday.get_or_none(Holiday.date == date_obj)
            if existing:
                existing.name = name
                existing.remark = remark
                existing.save()
            else:
                Holiday.create(date=date_obj, name=name, remark=remark)
            imported += 1

        except Exception as e:
            errors.append(f'第{i}行：{str(e)}')

    return imported, errors


def import_comprehensive(wb):
    """综合导入：一次性导入所有工作表的数据"""
    results = {
        'room': {'imported': 0, 'updated': 0, 'errors': []},
        'teacher': {'imported': 0, 'updated': 0, 'errors': []},
        'class': {'imported': 0, 'updated': 0, 'errors': []},
        'holiday': {'imported': 0, 'errors': []},
        'course': {'imported': 0, 'updated': 0, 'errors': []},
    }
    
    # 按顺序导入：教室 -> 教师 -> 班级 -> 节假日 -> 课程
    sheet_order = ['教室数据', '教师数据', '班级数据', '节假日数据', '课程数据']
    type_mapping = {
        '教室数据': 'room',
        '教师数据': 'teacher',
        '班级数据': 'class',
        '节假日数据': 'holiday',
        '课程数据': 'course',
    }
    
    for sheet_name in sheet_order:
        if sheet_name not in wb.sheetnames:
            continue
            
        ws = wb[sheet_name]
        data_type = type_mapping[sheet_name]
        
        headers = [cell.value for cell in ws[1]]
        data_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(cell is None for cell in row):
                continue
            data_rows.append(list(row))
        
        if not data_rows:
            continue
            
        try:
            if data_type == 'room':
                imported, updated, skipped, errors = import_rooms(data_rows, headers)
            elif data_type == 'teacher':
                imported, updated, skipped, errors = import_teachers(data_rows, headers)
            elif data_type == 'class':
                imported, updated, skipped, errors = import_classes(data_rows, headers)
            elif data_type == 'holiday':
                imported, errors = import_holidays(data_rows, headers)
            elif data_type == 'course':
                imported, updated, skipped, errors = import_courses(data_rows, headers)
            else:
                continue
                
            results[data_type]['imported'] = imported
            results[data_type]['updated'] = updated
            results[data_type]['errors'] = errors
        except Exception as e:
            results[data_type]['errors'] = [str(e)]
    
    # 计算总数
    total_imported = sum(r['imported'] for r in results.values())
    total_updated = sum(r['updated'] for r in results.values())
    all_errors = []
    for r in results.values():
        all_errors.extend(r['errors'])
    
    return jsonify({
        'count': total_imported + total_updated,
        'details': {
            'room': {'imported': results['room']['imported'], 'updated': results['room']['updated']},
            'teacher': {'imported': results['teacher']['imported'], 'updated': results['teacher']['updated']},
            'class': {'imported': results['class']['imported'], 'updated': results['class']['updated']},
            'holiday': {'imported': results['holiday']['imported']},
            'course': {'imported': results['course']['imported'], 'updated': results['course']['updated']},
        },
        'errors': '; '.join(all_errors[:10]) if all_errors else None
    }), 207 if all_errors else 200


def import_rooms(data_rows, headers):
    """导入教室数据"""
    imported = 0
    updated = 0
    skipped = 0
    errors = []

    def find_header(key):
        for h in headers:
            if h and key in str(h):
                return headers.index(h)
        return -1

    number_idx = find_header('编号') if find_header('编号') >= 0 else find_header('教室编号')
    name_idx = find_header('名称') if find_header('名称') >= 0 else 1
    capacity_idx = find_header('容量')
    type_idx = find_header('类型') if find_header('类型') >= 0 else 3

    for i, row in enumerate(data_rows, start=2):
        try:
            if len(row) < 4 or row[0] is None:
                continue

            room_number = str(row[number_idx]).strip() if number_idx >= 0 else ''
            name = str(row[name_idx]).strip() if name_idx >= 0 else ''
            capacity = int(row[capacity_idx]) if capacity_idx >= 0 and row[capacity_idx] else 0
            room_type = str(row[type_idx]).strip() if type_idx >= 0 else ''

            existing_room = Room.get_or_none(Room.room_number == room_number)
            if existing_room:
                existing_room.name = name
                existing_room.capacity = capacity
                existing_room.room_type = room_type
                existing_room.save()
                updated += 1
            else:
                Room.create(
                    room_number=room_number,
                    name=name,
                    capacity=capacity,
                    room_type=room_type,
                )
                imported += 1
        except Exception as e:
            errors.append(f'第{i}行：{str(e)}')

    return imported, updated, skipped, errors


def import_teachers(data_rows, headers):
    """导入教师数据"""
    imported = 0
    updated = 0
    skipped = 0
    errors = []

    def find_header(key):
        for h in headers:
            if h and key in str(h):
                return headers.index(h)
        return -1

    number_idx = find_header('工号') if find_header('工号') >= 0 else 0
    name_idx = find_header('姓名') if find_header('姓名') >= 0 else 1
    courses_idx = find_header('课程')
    max_idx = find_header('课次') if find_header('课次') >= 0 else 3

    for i, row in enumerate(data_rows, start=2):
        try:
            if len(row) < 4 or row[0] is None:
                continue

            teacher_number = str(row[number_idx]).strip() if number_idx >= 0 else ''
            name = str(row[name_idx]).strip() if name_idx >= 0 else ''
            courses = str(row[courses_idx]).strip() if courses_idx >= 0 and row[courses_idx] else ''
            max_weekly = int(row[max_idx]) if max_idx >= 0 and row[max_idx] else 5
            
            course_list = [c.strip() for c in courses.split(',') if c.strip()]
            teachable_courses_json = json.dumps(course_list)

            existing_teacher = Teacher.get_or_none(Teacher.teacher_number == teacher_number)
            if existing_teacher:
                existing_teacher.name = name
                existing_teacher.teachable_courses = teachable_courses_json
                existing_teacher.max_weekly_sessions = max_weekly
                existing_teacher.save()
                updated += 1
            else:
                Teacher.create(
                    teacher_number=teacher_number,
                    name=name,
                    teachable_courses=teachable_courses_json,
                    max_weekly_sessions=max_weekly,
                )
                imported += 1
        except Exception as e:
            errors.append(f'第{i}行：{str(e)}')

    return imported, updated, skipped, errors


def import_classes(data_rows, headers):
    """导入班级数据"""
    imported = 0
    updated = 0
    skipped = 0
    errors = []

    def find_header(key):
        for h in headers:
            if h and key in str(h):
                return headers.index(h)
        return -1

    number_idx = find_header('编号') if find_header('编号') >= 0 else 0
    name_idx = find_header('名称') if find_header('名称') >= 0 else 1
    count_idx = find_header('人数') if find_header('人数') >= 0 else 2
    dept_idx = find_header('院系') if find_header('院系') >= 0 else 3

    for i, row in enumerate(data_rows, start=2):
        try:
            if len(row) < 4 or row[0] is None:
                continue

            class_number = str(row[number_idx]).strip() if number_idx >= 0 else ''
            name = str(row[name_idx]).strip() if name_idx >= 0 else ''
            student_count = int(row[count_idx]) if count_idx >= 0 and row[count_idx] else 0
            department = str(row[dept_idx]).strip() if dept_idx >= 0 else ''

            existing_class = SchoolClass.get_or_none(SchoolClass.class_number == class_number)
            if existing_class:
                existing_class.name = name
                existing_class.student_count = student_count
                existing_class.department = department
                existing_class.save()
                updated += 1
            else:
                SchoolClass.create(
                    class_number=class_number,
                    name=name,
                    student_count=student_count,
                    department=department,
                )
                imported += 1
        except Exception as e:
            errors.append(f'第{i}行：{str(e)}')

    return imported, updated, skipped, errors


def import_courses(data_rows, headers):
    """导入课程数据"""
    imported = 0
    updated = 0
    skipped = 0
    errors = []

    def find_header(key):
        for h in headers:
            if h and key in str(h):
                return headers.index(h)
        return -1

    number_idx = find_header('课程编号') if find_header('课程编号') >= 0 else 0
    name_idx = find_header('课程名称') if find_header('课程名称') >= 0 else 1
    type_idx = find_header('课程类型') if find_header('课程类型') >= 0 else 2
    hours_idx = find_header('课时') if find_header('课时') >= 0 else 3
    teacher_name_idx = find_header('教师')
    class_name_idx = find_header('班级')

    for i, row in enumerate(data_rows, start=2):
        try:
            if len(row) < 6 or row[0] is None:
                continue

            course_number = str(row[number_idx]).strip() if number_idx >= 0 else ''
            name = str(row[name_idx]).strip() if name_idx >= 0 else ''
            course_type = str(row[type_idx]).strip() if type_idx >= 0 else ''
            total_hours = int(row[hours_idx]) if hours_idx >= 0 and row[hours_idx] else 64
            teacher_name = str(row[teacher_name_idx]).strip() if teacher_name_idx >= 0 and row[teacher_name_idx] else ''
            class_name = str(row[class_name_idx]).strip() if class_name_idx >= 0 and row[class_name_idx] else ''

            teacher = Teacher.get_or_none(Teacher.name == teacher_name)
            if not teacher:
                errors.append(f'第{i}行：找不到教师"{teacher_name}"')
                continue

            school_class = SchoolClass.get_or_none(SchoolClass.name == class_name)
            if not school_class:
                errors.append(f'第{i}行：找不到班级"{class_name}"')
                continue

            existing_course = Course.get_or_none(Course.course_number == course_number)
            if existing_course:
                existing_course.name = name
                existing_course.course_type = course_type
                existing_course.total_hours = total_hours
                existing_course.teacher = teacher
                existing_course.school_class = school_class
                existing_course.save()
                course = existing_course
                updated += 1
            else:
                course = Course.create(
                    course_number=course_number,
                    name=name,
                    course_type=course_type,
                    total_hours=total_hours,
                    teacher=teacher,
                    school_class=school_class,
                )
                imported += 1

            # 创建教学班记录（用于排课，assigned_day/period由排课算法分配）
            existing_tc = TeachingClass.get_or_none(
                (TeachingClass.course == course) &
                (TeachingClass.school_class == school_class) &
                (TeachingClass.teacher == teacher)
            )
            if not existing_tc:
                TeachingClass.create(
                    course=course,
                    school_class=school_class,
                    teacher=teacher,
                )
        except Exception as e:
            errors.append(f'第{i}行：{str(e)}')

    return imported, updated, skipped, errors
